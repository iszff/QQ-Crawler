#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import selenium
from selenium import webdriver
import time
import requests
import re
import openpyxl
from openpyxl import load_workbook


def login_in(url, qq, password):
    driver = webdriver.Chrome()
    driver.set_page_load_timeout(100)
    driver.get(url)
    time.sleep(6)
    driver.switch_to.frame('login_frame')
    driver.find_element_by_xpath('//*[@id="switcher_plogin"]').click()
    driver.find_element_by_xpath('//*[@id="u"]').clear()
    driver.find_element_by_xpath('//*[@id="u"]').send_keys(qq)
    driver.find_element_by_xpath('//*[@id="p"]').clear()
    driver.find_element_by_xpath('//*[@id="p"]').send_keys(password)
    driver.find_element_by_xpath('//*[@id="login_button"]').click()
    time.sleep(6)
    
    html = driver.page_source
    fw = re.findall(r'主人设置了权限，您可通过以下方式访问',html)
    if len(fw)!=0:
        driver.delete_all_cookies()
        driver.quit()
        return 0,0,0
    
    else:  
        cookies = driver.get_cookies()
        cookie = {}
        for elem in cookies:
       	    cookie[elem['name']] = elem['value']
        driver.delete_all_cookies()    

        hashes=5381
        for letter in cookie['p_skey']:
            hashes += (hashes << 5) + ord(letter)
        gtk=str(hashes & 0x7fffffff)   
	    
        g_qzonetoken = re.findall(r'window\.g\_qzonetoken \= \(function\(\)\{ try\{return \".*?\"',html)
        qzonetoken=g_qzonetoken[0].split('"')[-2]
	    
        driver.quit()
	    
        return cookie, gtk, qzonetoken   


def enter_qzone(friend_qq,friend_name,my_qq,my_password,pub_per,count,save_path):
       
    url = 'https://user.qzone.qq.com/'+friend_qq+'/311'
    (cookie, gtk, qzonetoken) = login_in(url, my_qq, my_password)

    if cookie == 0:
        wb = load_workbook(save_path)
        ws = wb.get_sheet_by_name("Black")
        wtRow = str(ws.max_row + 1)
        ws['A' + wtRow].value = friend_name
        ws['B' + wtRow].value = friend_qq
        ws['C' + wtRow].value = count

    else:
        urlOrd = 'https://user.qzone.qq.com/proxy/domain/taotao.qq.com/cgi-bin/emotion_cgi_msglist_v6?uin='
        sec='&ftype=0&sort=0&pos='
        third='&num=20&replynum=100&g_tk='
        forth='&callback=_preloadCallback&code_version=1&format=jsonp&need_private_comment=1&qzonetoken='
        fifth='&g_tk='
    
        wb = load_workbook(save_path)
        sheetName = str(count) + '_' + friend_qq + '_' + friend_name
        ws = wb.create_sheet(sheetName)
        wc = ws['A1':'O1'][0] 
        sheetHead = ["名字", "号码", "内容", "字数", "日期", "时刻", "转发数", "评论数","评论内容", "pos_x", "pos_y", "图片数", "说说id", "手机", "是否原创"]
    
        for i in range(15):
            wc[i].value = sheetHead[i]

        for ye in range(400):
            pos = str(ye*20)
            emotionurl=urlOrd+ friend_qq + sec + pos + third + gtk + forth + qzonetoken+ fifth + gtk
            textpage = get_emotion(emotionurl, cookie)
            time.sleep(10)

            if re.findall(r'"msglist":.*?,"name"',textpage)==['"msglist":null,"name"']:
                break
            else:
                (num, Info) = process(textpage, pub_per, friend_qq, friend_name)
                if num==0:
                    break
                else:
                    save_pageInfo(ws, num, Info)
                    wb.save(save_path)
    wb.save(save_path)        


def get_emotion(url,cookie):
    header= {'accept-encoding': 'gzip, deflate, br', 'accept-language': 'zh-CN,zh;q=0.8', 'host': 'h5.qzone.qq.com', 'user-agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.115 Safari/537.36', 'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8', 'connection': 'keep-alive'}
    r = requests.get(url, headers=header, cookies=cookie) 
    return r.text
       
    
def process(ss, time, qqnum, qqname):
    str_withRt = '\"certified\"\:.*?"createTime":"'+ time +'.*?"\,\"created_time\".*?\"wbid\"'
    ssWithRt = re.findall(str_withRt, ss)

    if len(ssWithRt)==0:
        return 0, 'null'
    else:
        str_withoutRt = '\"certified\"\:.*?"createTime":"' + time + '.*?"\,\"created_time\".*?\"right\"\:'
        ssWithoutRt = re.findall(str_withoutRt, ss)
        ssNum = len(ssWithoutRt)

        pageInfo=[]
        
        for i in range(ssNum):

            ssInfo=[qqname,qqnum]

            elem = ssWithoutRt[i]

            conlist = re.findall(r'\"conlist\".*?\"content\"\:',elem)[0]
            if conlist == '"conlist":null,"content":':
                words = ''
            else:
                try:
                    words = conlist.split('"conlist":[{"con":"')[1].split('","type":')[0]
                except:
                    words = ''
            ssInfo.append(words)
            wdl = len(words)   
            ssInfo.append(wdl)
            
            try: 
               postTime = re.findall('\"createTime\"\:.{14}\"created_time\":\d.*?\,\"editMask\"',elem)[0]
               day = postTime.split('"createTime":"')[1].split('","created_time"')[0]
            except:
               day = '2018年'
            ssInfo.append(day)

            h = postTime.split(',"created_time":')[1].split(',"editMask"')[0]
            hour = ( int(h) - 1514649600 ) %86400 /3600
            ssInfo.append(hour) 
            
            fwd = re.findall('\"fwdnum":.*?\,',elem)[0].split('"fwdnum":')[1].split(',')[0] 
            ssInfo.append(int(fwd)) 
          
            cm = re.findall('\"cmtnum\"\:.*?\,',elem)[0].split('"cmtnum":')[1].split(',')[0]
            cmnum = int(cm)
            ssInfo.append(cmnum)
            comment = ''
            if cmnum==0:
                comment = ''
            else:
                com =re.findall('\"abledel\"\:.*?\,\"content\"\:\".*?\"',elem)
                for j in com:
                    b = re.split('\"abledel\"\:\d\,\"content\"\:\"',j)[1].split('"')[-2] + ''#Orz
                    comment += b
            ssInfo.append(comment)
            
            pos_x = re.findall('\"pos\_x\"\:\".*?\"',elem)[0].split('"pos_x":"')[1].split('"')[0]
            pos_y = re.findall('\"pos\_y\"\:\".*?\"',elem)[0].split('"pos_y":"')[1].split('"')[0]
            ssInfo.append(pos_x)
            ssInfo.append(pos_y)

            elem1 = ssWithRt[i]                                         
           
            pictotal = re.findall('"pictotal":\d',elem1)
            if pictotal == []:
                picnum = 0
            else:
                picnum = int(pictotal[0].split('"pictotal":')[1])
            ssInfo.append(picnum)

            tid= re.findall('\"tid\"\:\".*?\"',elem1)[0].split('"tid":"')[1].split('"')[0]
            ssInfo.append(tid)

            phone = re.findall('\"source\_appid\"\:.*?\,\"source\_name\"\:.*?\"source\_url\"',elem1)[0].split('"source_appid":"","source_name":"')[1].split('","source_url"')[0]
            ssInfo.append(phone)
 
            rt = re.findall('\"rt\_con\"',elem1)
            if rt==[]:
                ori = 1
            else:
                ori = 0
            ssInfo.append(ori)

            pageInfo.append(ssInfo)

        return ssNum, pageInfo
           
 
def save_pageInfo(ws, n, info):
    readyRow = ws.max_row + 1
    for i in range(n):
        per_ssInfo = info[i]

        wtRow = str(readyRow + i)
        cellBegin = 'A' + wtRow
        cellEnd = 'O' + wtRow
        wc = ws[cellBegin:cellEnd][0]
        for j in range(15):
            wc[j].value = per_ssInfo[j]

    return


def main():
    start = time.time()
    save_path = '你的保存路径'
    my_qq = '你的qq'
    my_password = '你的密码'
    pub_per = '2018年'
    wb = load_workbook(save_path)
    ws = wb.get_sheet_by_name("friend_info")
    wc_name=ws['A']
    wc_qq=ws['B']
    total = len(wc_name)-1
    Name = []
    QQ = []
    
    for i in range(total):
        n = wc_name[i+1].value
        Name.append(n) 
        q = wc_qq[i+1].value
        QQ.append(q)
    wb.close()
    
    for j in range(total):
        count = j+1
        print('\r当前进度:{:.2f}%'.format(count*100/total),end=' ')
        friend_name =Name[j]
        friend_qq = QQ[j]
        enter_qzone(friend_qq, friend_name, my_qq, my_password, pub_per, count,save_path)
        time.sleep(8)
    end = time.time()
    T = end-start
    print('用时:{:f}s'.format(T))

main()
