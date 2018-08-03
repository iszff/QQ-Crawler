#!/usr/bin/env python3
# -*-coding: utf-8 -*-
import selenium 
from selenium import webdriver
import requests
import openpyxl
from openpyxl import load_workbook
import time
import re 

def login_in(url, qq, password):
    driver = webdriver.Chrome()
    driver.get(url)
    time.sleep(6)
    driver.switch_to.frame('login_frame')
    driver.find_element_by_xpath('//*[@id="switcher_plogin"]').click()
    driver.find_element_by_xpath('//*[@id="u"]').clear()
    driver.find_element_by_xpath('//*[@id="u"]').send_keys(qq)
    driver.find_element_by_xpath('//*[@id="p"]').clear()
    driver.find_element_by_xpath('//*[@id="p"]').send_keys(password)
    driver.find_element_by_xpath('//*[@id="login_button"]').click()
    time.sleep(6)

    cookies = driver.get_cookies()
    cookie = {}
    for elem in cookies:
        cookie[elem['name']] = elem['value']
 
    hashes=5381
    for letter in cookie['p_skey']:
        hashes += (hashes << 5) + ord(letter)
    gtk=str(hashes & 0x7fffffff)

    html = driver.page_source 
    g_qzonetoken = re.findall(r'window\.g\_qzonetoken \= \(function\(\)\{ try\{return \".*?\"',html)
    qzonetoken=g_qzonetoken[0].split('"')[-2]

    driver.quit()

    return cookie, gtk, qzonetoken


def enter_qzone(my_qq, my_password, save_path):
    url = 'https://user.qzone.qq.com/'+my_qq+'/311'
    (cookie, gtk, qzonetoken) = login_in(url, my_qq, my_password)

    urlOrd = 'https://user.qzone.qq.com/proxy/domain/r.qzone.qq.com/cgi-bin/tfriend/friend_hat_get.cgi?hat_seed=1&uin='
    second ='&fupdate=1&g_tk='
    third = '&qzonetoken='
    fourth = '&g_tk='
    haturl= urlOrd + my_qq + second + gtk + third + qzonetoken + fourth + gtk

    textpage = get_friendhat(haturl, cookie)
    process_save(textpage,save_path)


def get_friendhat(url,cookie):
    header= {'accept-encoding': 'gzip, deflate, br', 'accept-language': 'zh-CN,zh;q=0.8', 'host': 'h5.qzone.qq.com', 'user-agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.115 Safari/537.36', 'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8', 'connection': 'keep-alive'}
    r = requests.get(url, headers=header, cookies=cookie)
    return r.text

def process_save(text,save_path):
    friend=re.findall('\".*?\"\:\{\n\"realname\"\:\".*?\"\}',text)
    frNum = len(friend)

    Name = []
    for elem in friend:
        n =elem.split('"realname":"')[1].split('"}')[0]
        Name.append(n)
    QQ = []
    for elem in friend:
        q = elem.split('"')[1]
        QQ.append(q)

    wb = load_workbook(save_path)
    ws = wb.get_sheet_by_name("friend_info")
    wc_name=ws['A2':'A'+str(frNum+2)]
    wc_qq=ws['B2':'B'+str(frNum+2)]
    for i in range(frNum):
        wc_name[i][0].value = Name[i]
        wc_qq[i][0].value = QQ[i]

    wb.save(save_path)

def main():
    my_qq = '你的qq'
    my_password = '你的密码'
    save_path = '你的保存路径'
    enter_qzone(my_qq, my_password, save_path)

main()
